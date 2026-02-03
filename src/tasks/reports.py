import os
import tempfile
from datetime import datetime, timedelta

from celery import Task

from src.celery_app import celery_app
from src.database import AsyncSessionLocal
from src.repositories.batch import BatchRepository
from src.repositories.product import ProductRepository
from src.services.minio_service import minio_service


@celery_app.task(bind=True, max_retries=3)
def generate_batch_report(
    self: Task, batch_id: int, format: str = "excel", user_email: str | None = None
) -> dict:
    """
    Генерация детального отчета по партии.

    Args:
        batch_id: ID партии
        format: "excel" или "pdf"
        user_email: Email для отправки уведомления (опционально)

    Returns:
        {
            "success": True,
            "file_url": "...",
            "file_name": "...",
            "file_size": 152400,
            "expires_at": "2024-02-07T00:00:00Z"
        }
    """
    import asyncio

    async def _generate():
        async with AsyncSessionLocal() as session:
            # Get batch with products
            batch_repo = BatchRepository(session)
            batch = await batch_repo.get_by_id(batch_id, with_products=True)

            if not batch:
                return {"success": False, "error": f"Batch {batch_id} not found"}

            # Get statistics
            product_repo = ProductRepository(session)
            stats = await product_repo.get_statistics(batch_id)

            # Generate file
            if format == "excel":
                file_path = _generate_excel_report(batch, stats)
            elif format == "pdf":
                file_path = _generate_pdf_report(batch, stats)
            else:
                return {"success": False, "error": f"Unsupported format: {format}"}

            # Upload to MinIO
            file_name = (
                f"batch_{batch_id}_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{format}"
            )
            file_url = minio_service.upload_file(
                bucket="reports",
                file_path=file_path,
                object_name=file_name,
                expires_days=7,
            )

            file_size = os.path.getsize(file_path)

            expires_at = (datetime.utcnow() + timedelta(days=7)).isoformat() + "Z"

            # Send webhook event
            from src.repositories.webhook import WebhookRepository
            from src.services.webhook_service import webhook_service
            from src.tasks.webhooks import send_webhook_delivery

            webhook_repo = WebhookRepository(session)
            subscriptions = await webhook_repo.get_active_subscriptions_for_event(
                "report_generated"
            )

            for subscription in subscriptions:
                payload = webhook_service.create_webhook_payload(
                    "report_generated",
                    {
                        "batch_id": batch_id,
                        "report_type": format,
                        "file_url": file_url,
                        "file_name": file_name,
                        "file_size": file_size,
                        "expires_at": expires_at,
                    },
                )
                delivery = await webhook_repo.create_delivery(
                    subscription_id=subscription.id,
                    event_type="report_generated",
                    payload=payload,
                )
                send_webhook_delivery.delay(delivery.id)

            # Cleanup temp file
            try:
                os.remove(file_path)
            except Exception:
                pass

            # Send email notification if provided
            if user_email:
                try:
                    import logging

                    logger = logging.getLogger(__name__)
                    logger.info(
                        f"Report generated for batch {batch_id}, email: {user_email}, file: {file_url}"
                    )
                except Exception as e:
                    # Don't fail the task if email fails
                    print(f"Failed to send email notification: {e}")

            return {
                "success": True,
                "file_url": file_url,
                "file_name": file_name,
                "file_size": file_size,
                "expires_at": expires_at,
                "email_sent": user_email is not None,
            }

    try:
        result = asyncio.run(_generate())
        return result
    except Exception as exc:
        raise self.retry(exc=exc, countdown=2**self.request.retries) from None


def _generate_excel_report(batch, stats) -> str:
    """Generate Excel report"""
    from openpyxl import Workbook

    wb = Workbook()

    # Sheet 1: Batch Info
    ws1 = wb.active
    ws1.title = "Информация о партии"

    ws1.append(["Номер партии:", batch.batch_number])
    ws1.append(["Дата партии:", str(batch.batch_date)])
    ws1.append(["Статус:", "Закрыта" if batch.is_closed else "Открыта"])
    ws1.append(["Рабочий центр:", batch.work_center.name if batch.work_center else ""])
    ws1.append(["Смена:", batch.shift])
    ws1.append(["Бригада:", batch.team])
    ws1.append(["Номенклатура:", batch.nomenclature])
    ws1.append(["Код ЕКН:", batch.ekn_code])
    ws1.append(["Начало смены:", batch.shift_start.strftime("%Y-%m-%d %H:%M:%S")])
    ws1.append(["Окончание смены:", batch.shift_end.strftime("%Y-%m-%d %H:%M:%S")])

    # Sheet 2: Products
    ws2 = wb.create_sheet("Продукция")
    ws2.append(["ID", "Уникальный код", "Аггрегирована", "Дата аггрегации"])

    for product in batch.products:
        ws2.append(
            [
                product.id,
                product.unique_code,
                "Да" if product.is_aggregated else "Нет",
                product.aggregated_at.strftime("%Y-%m-%d %H:%M:%S")
                if product.aggregated_at
                else "-",
            ]
        )

    # Sheet 3: Statistics
    ws3 = wb.create_sheet("Статистика")
    ws3.append(["Всего продукции:", stats["total_products"]])
    ws3.append(["Аггрегировано:", stats["aggregated"]])
    ws3.append(["Осталось:", stats["remaining"]])
    ws3.append(["Процент выполнения:", f"{stats['aggregation_rate']:.2f}%"])

    # Save to temp file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    temp_file.close()

    return temp_file.name


def _generate_pdf_report(batch, stats) -> str:
    """Generate PDF report"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()

    # Title
    story.append(Paragraph("Отчет по партии", styles["Title"]))
    story.append(Spacer(1, 12))

    # Batch Info
    data = [
        ["Номер партии:", str(batch.batch_number)],
        ["Дата партии:", str(batch.batch_date)],
        ["Статус:", "Закрыта" if batch.is_closed else "Открыта"],
        ["Рабочий центр:", batch.work_center.name if batch.work_center else ""],
        ["Смена:", batch.shift],
        ["Бригада:", batch.team],
        ["Номенклатура:", batch.nomenclature],
    ]

    table = Table(data, colWidths=[150, 300])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.grey),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ("BACKGROUND", (1, 0), (1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    story.append(table)
    story.append(Spacer(1, 12))

    # Statistics
    story.append(Paragraph("Статистика", styles["Heading2"]))
    stats_data = [
        ["Всего продукции:", str(stats["total_products"])],
        ["Аггрегировано:", str(stats["aggregated"])],
        ["Осталось:", str(stats["remaining"])],
        ["Процент выполнения:", f"{stats['aggregation_rate']:.2f}%"],
    ]

    stats_table = Table(stats_data, colWidths=[150, 300])
    stats_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.lightblue),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    story.append(stats_table)

    doc.build(story)
    temp_file.close()

    return temp_file.name
